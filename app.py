from flask import Flask, request, send_file, render_template_string
import pandas as pd
import zipfile
from io import BytesIO

# 👉 IMPORT MAPPER
from mapper import process_sheet

app = Flask(__name__)

app.config['MAX_CONTENT_LENGTH'] = 50 * 1024 * 1024

uploaded_file = None

@app.route('/', methods=['GET', 'POST'])
def upload():
    global uploaded_file

    if request.method == 'POST':
        file = request.files['file']

        uploaded_file = BytesIO(file.read())
        uploaded_file.seek(0)

        xls = pd.ExcelFile(uploaded_file)
        sheets = xls.sheet_names

        return render_template_string('''
        <h2>Select Sheets</h2>
        <form method="POST" action="/process">
            {% for s in sheets %}
                <input type="checkbox" name="sheets" value="{{s}}" checked> {{s}}<br>
            {% endfor %}
            <br>
            <button type="submit">Process</button>
        </form>
        ''', sheets=sheets)

    return '''
    <h2>Upload Excel</h2>
    <form method="POST" enctype="multipart/form-data">
        <input type="file" name="file" required>
        <button type="submit">Upload</button>
    </form>
    '''

@app.route('/process', methods=['POST'])
def process():
    global uploaded_file

    uploaded_file.seek(0)
    xls = pd.ExcelFile(uploaded_file)

    selected_sheets = request.form.getlist('sheets')
    output_files = []

    try:
        gst_df = pd.read_excel(xls, sheet_name="GST", header=None)
    except:
        gst_df = pd.DataFrame()

    for sheet in selected_sheets:
        try:
            if sheet == "GST":
                continue

            df = pd.read_excel(xls, sheet_name=sheet, header=None)
            out_df = process_sheet(df, gst_df)

            if out_df is None or out_df.empty:
                continue

            # =========================
            # ✅ UPDATED EXPORT WITH FORMATTING
            # =========================
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

            output = BytesIO()

            with pd.ExcelWriter(output, engine="openpyxl") as writer:
                out_df.to_excel(writer, index=False, sheet_name="Sheet1")

                ws = writer.sheets["Sheet1"]

                # Header style
                header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

                for cell in ws[1]:
                    cell.font = Font(bold=True)
                    cell.fill = header_fill
                    cell.alignment = Alignment(horizontal="center", vertical="center")

                # Auto column width
                for col in ws.columns:
                    max_length = 0
                    col_letter = col[0].column_letter

                    for cell in col:
                        try:
                            if cell.value:
                                max_length = max(max_length, len(str(cell.value)))
                        except:
                            pass

                    ws.column_dimensions[col_letter].width = max_length + 2

                # Borders + alignment
                thin = Border(
                    left=Side(style='thin'),
                    right=Side(style='thin'),
                    top=Side(style='thin'),
                    bottom=Side(style='thin')
                )

                for row in ws.iter_rows():
                    for cell in row:
                        cell.border = thin
                        cell.alignment = Alignment(vertical="center")

            output.seek(0)
            output_files.append((f"{sheet}.xlsx", output))

        except Exception as e:
            print("ERROR:", sheet, e)
            continue

    if not output_files:
        return "No output generated"

    memory_file = BytesIO()

    with zipfile.ZipFile(memory_file, 'w') as zf:
        for filename, data in output_files:
            zf.writestr(filename, data.getvalue())

    memory_file.seek(0)

    return send_file(memory_file, download_name="output.zip", as_attachment=True)


if __name__ == "__main__":
    app.run(debug=True)
