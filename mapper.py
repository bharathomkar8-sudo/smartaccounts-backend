import pandas as pd

COLUMNS = [
    "Voucher Type","VCH No / Inv No","Description","VCH Date","Order No","Order Date","Other Ref","POS",
    "Party Name","Address","State","Pincode","Party GSTIN",
    "Consignee Name","Con Address","Consignee State","Consignee Pincode","Con GSTIN",
    "Item Name / Code","Is Item Header","width","Height","Qty","Extraudf","Billedqty","Rate",
    "Taxable Value","Dis%","Amount","Sales Ledger",
    "CGST Ledger","CGST Amt","SGST Ledger","SGST Amount",
    "IGST Ledger","IGST Amount","Round off","Invoice Amt","Item header"
]

def clean(val):
    if pd.isna(val):
        return ""
    val = str(val).strip()
    if val.lower() == "nan":
        return ""
    return val

# ✅ DATE FIX
def format_date(val):
    try:
        return pd.to_datetime(val, dayfirst=True).strftime("%d-%m-%Y")
    except:
        return ""

def process_sheet(df, gst_df):

    rows = []

    voucher_type = "Sales E-Invoice"
    vch_no = clean(df.iloc[10, 16])     
    vch_date = format_date(df.iloc[11, 16])
    order_no = clean(df.iloc[19, 1])
    order_date = format_date(df.iloc[20, 1])   # ✅ FIXED
    other_ref = clean(df.iloc[12, 16])
    pos = clean(df.iloc[14, 5])

    state = clean(df.iloc[14, 1])
    pincode = clean(df.iloc[15, 1])
    gst = clean(df.iloc[16, 1])

    party_name = ""

    try:
        match = gst_df[gst_df.iloc[:, 0] == gst]
        if not match.empty:
            party_name = match.iloc[0, 1]
        else:
            party_name = clean(df.iloc[10, 0])
    except:
        party_name = clean(df.iloc[10, 0])

    address_lines = [
        clean(df.iloc[11, 0]),
        clean(df.iloc[12, 0]),
        clean(df.iloc[13, 0])
    ]

    con_address_lines = [
        clean(df.iloc[12, 4]),
        clean(df.iloc[13, 4])
    ]

    for i in range(25, len(df)):

        desc = df.iloc[i, 1]
        desc = clean(desc)

        if desc.lower() == "end here":
            break

        row = dict.fromkeys(COLUMNS, "")

        row["Voucher Type"] = voucher_type
        row["VCH No / Inv No"] = vch_no
        row["VCH Date"] = vch_date
        row["Order No"] = order_no
        row["Order Date"] = order_date
        row["Other Ref"] = other_ref
        row["POS"] = pos

        row["State"] = state
        row["Pincode"] = pincode
        row["Party GSTIN"] = gst

        row["Party Name"] = party_name
        row["Consignee Name"] = party_name

        row["Consignee State"] = clean(df.iloc[14, 5])
        row["Consignee Pincode"] = clean(df.iloc[15, 5])
        row["Con GSTIN"] = gst

        row["Description"] = desc
        row["Item header"] = desc

        if i - 25 < len(address_lines):
            row["Address"] = address_lines[i - 25]

        if i - 25 < len(con_address_lines):
            row["Con Address"] = con_address_lines[i - 25]

        item_val = df.iloc[i, 2]

        if isinstance(item_val, (int, float)) and item_val > 1:
            row["Item Name / Code"] = str(int(item_val))
        else:
            row["Item Name / Code"] = "Header"

        if row["Item Name / Code"] == "Header":
            row["Is Item Header"] = "Yes"

        def safe(val):
            if pd.isna(val):
                return ""
            try:
                if float(val) == 0:
                    return ""
            except:
                pass
            return val

        row["width"] = safe(df.iloc[i, 3])
        row["Height"] = safe(df.iloc[i, 4])
        row["Qty"] = safe(df.iloc[i, 5])
        row["Extraudf"] = safe(df.iloc[i, 6])
        row["Billedqty"] = safe(df.iloc[i, 6])
        row["Rate"] = safe(df.iloc[i, 8])

        row["Dis%"] = safe(df.iloc[i, 9])

        try:
            billedqty = float(row["Billedqty"]) if row["Billedqty"] != "" else 0
            rate = float(row["Rate"]) if row["Rate"] != "" else 0
            dis = float(row["Dis%"]) if row["Dis%"] != "" else 0
        except:
            billedqty, rate, dis = 0, 0, 0

        taxable = round(billedqty * rate, 2)
        amount = round(taxable - (taxable * dis / 100), 2)

        row["Taxable Value"] = taxable if taxable else ""
        row["Amount"] = amount if amount else ""

        cgst_amt = safe(df.iloc[i, 12])
        sgst_amt = safe(df.iloc[i, 14])
        igst_amt = safe(df.iloc[i, 16])

        row["CGST Amt"] = cgst_amt
        row["SGST Amount"] = sgst_amt
        row["IGST Amount"] = igst_amt

        try:
            cgst_val = float(cgst_amt) if cgst_amt != "" else 0
            sgst_val = float(sgst_amt) if sgst_amt != "" else 0
            igst_val = float(igst_amt) if igst_amt != "" else 0
        except:
            cgst_val, sgst_val, igst_val = 0, 0, 0

        if cgst_val == 0 and igst_val == 0:
            row["Sales Ledger"] = "Header"
        elif igst_val > 0:
            row["Sales Ledger"] = "IGST Sales@18%"
        elif cgst_val > 0:
            row["Sales Ledger"] = "GST Sales@18%"

        row["CGST Ledger"] = "OUTPUT CGST" if cgst_val > 0 else ""
        row["SGST Ledger"] = "OUTPUT SGST" if sgst_val > 0 else ""
        row["IGST Ledger"] = "OUTPUT IGST" if igst_val > 0 else ""

        row["Round off"] = ""

        try:
            amt = float(row["Amount"]) if row["Amount"] != "" else 0
            cgst = float(cgst_amt) if cgst_amt != "" else 0
            sgst = float(sgst_amt) if sgst_amt != "" else 0
            igst = float(igst_amt) if igst_amt != "" else 0
        except:
            amt, cgst, sgst, igst = 0, 0, 0, 0

        invoice_total = amt + cgst + sgst + igst
        row["Invoice Amt"] = invoice_total if invoice_total else ""

        rows.append(row)

    df_out = pd.DataFrame(rows)

    # =========================
    # FORMATTING
    # =========================
    with pd.ExcelWriter("output.xlsx", engine="openpyxl") as writer:
        df_out.to_excel(writer, index=False)

        ws = writer.sheets["Sheet1"]

        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        # Header style
        header_fill = PatternFill(start_color="FFFF00", end_color="FFFF00", fill_type="solid")

        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")

        # Auto column width
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter

            for cell in col:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass

            ws.column_dimensions[col_letter].width = max_length + 2

        # Borders
        thin = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in ws.iter_rows():
            for cell in row:
                cell.border = thin
                cell.alignment = Alignment(vertical="center")

    return df_out
